import pandas as pd
import numpy as np
import json
import os
import requests
import datetime
import re  # Add this import for regex
from tqdm import tqdm
from openpyxl import Workbook  # Add this import for Workbook

def download_file():
    """Download the Excel file from Google Drive"""
    file_id = "1eFXVNBA3w52zG-XPAWyMSN5niaj_SLKt"
    output_path = os.path.join(os.path.dirname(__file__), "downloaded_historic.xlsx")
    
    print("-- Iniciando -- Download do arquivo Excel")
    
    try:
        # First, get the confirmation token
        url = f"https://drive.google.com/uc?export=download&id={file_id}"
        session = requests.Session()
        response = session.get(url, stream=True)
        
        # Check if we need to confirm download
        token = None
        for key, value in response.cookies.items():
            if key.startswith('download_warning'):
                token = value
                break
        
        if token:
            url = f"https://drive.google.com/uc?export=download&confirm={token}&id={file_id}"
        else:
            # Try alternative method for large files
            url = f"https://drive.google.com/uc?export=download&id={file_id}&confirm=t"
        
        # Download the file
        print(f"Baixando arquivo para: {output_path}")
        response = session.get(url, stream=True)
        
        # Check if we got HTML instead of the file (happens with large files)
        content_type = response.headers.get('content-type', '')
        if 'text/html' in content_type.lower():
            print("Recebido HTML em vez do arquivo. Tentando método alternativo...")
            
            # Extract the download link from the HTML
            html_content = response.text
            download_link_match = None
            
            # Look for the download link in the HTML
            if "Google Drive can't scan this file for viruses" in html_content:
                # Extract the form action URL and parameters
                form_action_match = re.search(r'action="([^"]+)"', html_content)
                confirm_match = re.search(r'name="confirm" value="([^"]+)"', html_content)
                
                if form_action_match and confirm_match:
                    form_action = form_action_match.group(1)
                    confirm = confirm_match.group(1)
                    url = f"{form_action}?id={file_id}&export=download&confirm={confirm}"
                    response = session.get(url, stream=True)
            else:
                print("Não foi possível encontrar o link de download no HTML.")
                return None
        
        # Get file size for progress bar
        file_size = int(response.headers.get('content-length', 0))
        block_size = 1024  # 1 Kibibyte
        
        # Show download progress
        with open(output_path, 'wb') as f, tqdm(
                desc="Progresso",
                total=file_size,
                unit='iB',
                unit_scale=True,
                unit_divisor=1024,
            ) as bar:
            for data in response.iter_content(block_size):
                size = f.write(data)
                bar.update(size)
        
        print("Download completo.")
        
        # Verify file exists and has content
        if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
            print(f"Arquivo salvo em: {output_path}")
            
            # Check file type by reading first few bytes
            with open(output_path, 'rb') as f:
                header = f.read(8)
            
            # PK header indicates ZIP file (XLSX)
            if header.startswith(b'PK'):
                print("Arquivo detectado como XLSX (formato ZIP)")
                return output_path
            # Excel 97-2003 XLS format
            elif header.startswith(b'\xD0\xCF\x11\xE0'):
                print("Arquivo detectado como XLS (formato antigo)")
                return output_path
            else:
                print(f"Aviso: Formato de arquivo não reconhecido. Primeiros bytes: {header.hex()}")
                return output_path
        else:
            print("Erro: Arquivo não foi baixado corretamente.")
            return None
    except Exception as e:
        print(f"Erro ao baixar o arquivo: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def format_date(date_value):
    """Format date to dd/MM/yyyy"""
    if pd.isna(date_value):
        return None
    
    if isinstance(date_value, datetime.datetime) or isinstance(date_value, pd.Timestamp):
        return date_value.strftime("%d/%m/%Y")
    
    # Try to convert string to date
    try:
        if isinstance(date_value, str):
            if "/" in date_value:
                parts = date_value.split("/")
                if len(parts) == 3:
                    day, month, year = map(int, parts)
                    return f"{day:02d}/{month:02d}/{year}"
            elif "-" in date_value:
                date_obj = datetime.datetime.strptime(date_value, "%Y-%m-%d")
                return date_obj.strftime("%d/%m/%Y")
        
        # If it's a number (Excel serial date)
        if isinstance(date_value, (int, float)):
            # Convert Excel date to Python date
            date_obj = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(date_value))
            return date_obj.strftime("%d/%m/%Y")
    
    except Exception as e:
        print(f"Erro ao processar data: {date_value}, {str(e)}")
    
    return str(date_value)

# Add this function to check the Excel file structure
def inspect_excel_file(file_path):
    """Inspect the Excel file structure and formulas"""
    print("\n-- Inspecionando estrutura do arquivo Excel")
    
    try:
        # First try a direct pandas read to get a quick overview
        print("Tentando leitura rápida com pandas...")
        try:
            df_sample = pd.read_excel(file_path, nrows=10, header=None)
            print(f"Visão geral do arquivo (primeiras 10 linhas):")
            print(df_sample)
            
            # Check for date columns
            date_cols = []
            for col in df_sample.columns:
                if pd.api.types.is_datetime64_any_dtype(df_sample[col]):
                    date_cols.append(col)
            
            if date_cols:
                print(f"Colunas com datas detectadas: {date_cols}")
                
                # Show the most recent dates in these columns
                for col in date_cols:
                    non_na_dates = df_sample[col].dropna()
                    if len(non_na_dates) > 0:
                        latest_date = non_na_dates.max()
                        print(f"Data mais recente na coluna {col}: {latest_date}")
        except Exception as e:
            print(f"Erro na leitura rápida: {str(e)}")
        
        # Use openpyxl for more detailed inspection
        from openpyxl import load_workbook
        
        try:
            # Load workbook with data_only=True to get calculated values
            print("\nCarregando arquivo para inspeção detalhada...")
            wb = load_workbook(file_path, data_only=True, read_only=False)
            
            # Get the active sheet or first sheet
            if wb.active:
                sheet = wb.active
            elif len(wb.sheetnames) > 0:
                sheet = wb[wb.sheetnames[0]]
            else:
                print("Aviso: Não foi possível encontrar uma planilha ativa no arquivo.")
                return False
            
            print(f"Nome da planilha: {sheet.title}")
            
            # Get dimensions
            try:
                print(f"Dimensões: {sheet.dimensions}")
                max_row = sheet.max_row
                max_col = sheet.max_column
                print(f"Número de linhas: {max_row}, Número de colunas: {max_col}")
            except Exception as e:
                print(f"Erro ao obter dimensões: {str(e)}")
                max_row = 1000  # Assume a reasonable number
                max_col = 100
            
            # Look for the header row (first row with company names)
            header_row = None
            for row_idx in range(1, min(20, max_row)):
                row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(20, max_col), 4)]
                non_empty = [val for val in row_values if val is not None]
                if non_empty:
                    header_row = row_idx
                    print(f"Possível linha de cabeçalho encontrada: {row_idx}")
                    print(f"Valores: {non_empty}")
                    break
            
            if header_row is None:
                print("Aviso: Não foi possível encontrar uma linha de cabeçalho.")
                header_row = 1
            
            # Sample data from different columns to find dates
            print("\nProcurando por datas no arquivo...")
            latest_dates = {}
            
            # Check every 4 columns (company data blocks)
            for col_idx in range(1, min(100, max_col), 4):
                company_name = sheet.cell(row=header_row, column=col_idx).value
                if company_name:
                    print(f"\nAnalisando dados para: {company_name}")
                    
                    # Look for dates in this column
                    dates_found = []
                    for row_idx in range(header_row + 1, min(header_row + 100, max_row)):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        
                        # Try to identify if it's a date
                        if isinstance(cell_value, (datetime.datetime, datetime.date)):
                            dates_found.append(cell_value)
                        elif isinstance(cell_value, str) and ('/' in cell_value or '-' in cell_value):
                            try:
                                # Try to parse as date
                                if '/' in cell_value:
                                    parts = cell_value.split('/')
                                    if len(parts) == 3:
                                        # Assume day/month/year format
                                        day, month, year = map(int, parts)
                                        date_obj = datetime.datetime(year, month, day)
                                        dates_found.append(date_obj)
                                elif '-' in cell_value:
                                    date_obj = datetime.datetime.strptime(cell_value, "%Y-%m-%d")
                                    dates_found.append(date_obj)
                            except:
                                pass
                    
                    if dates_found:
                        latest_date = max(dates_found)
                        latest_dates[company_name] = latest_date
                        print(f"  Data mais recente: {latest_date.strftime('%d/%m/%Y')}")
                        
                        # Show a few sample dates
                        sample_dates = sorted(dates_found, reverse=True)[:5]
                        print(f"  Amostra de datas: {[d.strftime('%d/%m/%Y') for d in sample_dates]}")
            
            if latest_dates:
                overall_latest = max(latest_dates.values())
                print(f"\nData mais recente encontrada no arquivo: {overall_latest.strftime('%d/%m/%Y')}")
                print(f"Empresa com data mais recente: {max(latest_dates.items(), key=lambda x: x[1])[0]}")
            else:
                print("\nNenhuma data identificada no arquivo.")
            
            # Try to find the current date in the file
            today = datetime.datetime.now().date()
            yesterday = today - datetime.timedelta(days=1)
            print(f"\nVerificando se a data atual ({today.strftime('%d/%m/%Y')}) ou de ontem ({yesterday.strftime('%d/%m/%Y')}) está presente no arquivo...")
            
            today_found = False
            yesterday_found = False
            recent_dates = {}
            
            for col_idx in range(1, min(100, max_col), 4):
                company_name = sheet.cell(row=header_row, column=col_idx).value
                if not company_name:
                    continue
                    
                # Look for dates in this column more extensively
                for row_idx in range(header_row + 1, min(header_row + 500, max_row)):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    
                    # Handle both datetime and date objects correctly
                    if isinstance(cell_value, datetime.datetime):
                        date_value = cell_value.date()
                        if date_value == today:
                            print(f"Data atual encontrada para {company_name} na célula ({row_idx}, {col_idx})")
                            today_found = True
                        elif date_value == yesterday:
                            print(f"Data de ontem encontrada para {company_name} na célula ({row_idx}, {col_idx})")
                            yesterday_found = True
                            
                        # Track most recent date for each company
                        if company_name not in recent_dates or date_value > recent_dates[company_name]:
                            recent_dates[company_name] = date_value
                            
                    elif isinstance(cell_value, datetime.date):
                        if cell_value == today:
                            print(f"Data atual encontrada para {company_name} na célula ({row_idx}, {col_idx})")
                            today_found = True
                        elif cell_value == yesterday:
                            print(f"Data de ontem encontrada para {company_name} na célula ({row_idx}, {col_idx})")
                            yesterday_found = True
                            
                        # Track most recent date for each company
                        if company_name not in recent_dates or cell_value > recent_dates[company_name]:
                            recent_dates[company_name] = cell_value
            
            if not today_found and not yesterday_found:
                print("Nem a data atual nem a de ontem foram encontradas no arquivo.")
                
                # Show the most recent dates found for each company
                if recent_dates:
                    print("\nDatas mais recentes encontradas por empresa:")
                    sorted_companies = sorted(recent_dates.items(), key=lambda x: x[1], reverse=True)
                    
                    # Show top 10 most recent
                    for company, date in sorted_companies[:10]:
                        print(f"  {company}: {date.strftime('%d/%m/%Y')}")
                    
                    # Overall most recent
                    most_recent_company, most_recent_date = sorted_companies[0]
                    print(f"\nData mais recente no arquivo: {most_recent_date.strftime('%d/%m/%Y')} ({most_recent_company})")
                    
                    # Calculate how outdated the file is
                    days_outdated = (today - most_recent_date).days
                    print(f"O arquivo está desatualizado por {days_outdated} dias.")
                else:
                    print("Não foi possível identificar datas recentes no arquivo.")
                
                # Check if there are any dates from the current month
                current_month = today.month
                current_year = today.year
                current_month_dates = False
                
                for company, date in latest_dates.items():
                    if date.month == current_month and date.year == current_year:
                        print(f"Data do mês atual encontrada para {company}: {date.strftime('%d/%m/%Y')}")
                        current_month_dates = True
                
                if not current_month_dates:
                    print("Nenhuma data do mês atual encontrada no arquivo.")
                    print("O arquivo pode estar desatualizado.")
            
            return True
            
        except Exception as e:
            print(f"Erro na inspeção detalhada: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
            
    except Exception as e:
        print(f"Erro ao inspecionar arquivo Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

# After the inspect_excel_file function, add this new function:

def update_dates_in_file(file_path, output_path=None):
    """Update dates in the Excel file to current dates for testing purposes"""
    if output_path is None:
        # Create a new filename with _updated suffix
        base_name, ext = os.path.splitext(file_path)
        output_path = f"{base_name}_updated{ext}"
    
    print(f"\n-- Atualizando datas no arquivo Excel para testes")
    print(f"Arquivo original: {file_path}")
    print(f"Arquivo atualizado será salvo em: {output_path}")
    
    try:
        # Use openpyxl for more direct control over the file
        from openpyxl import load_workbook, Workbook
        
        print("Carregando arquivo com openpyxl...")
        # Load with data_only=True to get calculated values instead of formulas
        wb = load_workbook(file_path, data_only=True)
        
        if wb.active:
            sheet = wb.active
        elif len(wb.sheetnames) > 0:
            sheet = wb[wb.sheetnames[0]]
        else:
            print("Aviso: Não foi possível encontrar uma planilha ativa no arquivo.")
            return False
        
        print(f"Nome da planilha: {sheet.title}")
        max_row = sheet.max_row
        max_col = sheet.max_column
        print(f"Dimensões: {sheet.dimensions}")
        print(f"Número de linhas: {max_row}, Número de colunas: {max_col}")
        
        # Create a new workbook with static values
        print("Criando nova planilha com valores estáticos...")
        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.title = sheet.title
        
        # Copy all content from original to new workbook as static values
        print("Copiando conteúdo como valores estáticos...")
        for row_idx in range(1, max_row + 1):
            if row_idx % 100 == 0:
                print(f"  Copiando linha {row_idx} de {max_row}...")
            for col_idx in range(1, max_col + 1):
                try:
                    # Get the value (not the formula)
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    # Set the value in the new sheet
                    new_sheet.cell(row=row_idx, column=col_idx).value = cell_value
                except Exception as e:
                    print(f"Erro ao copiar célula ({row_idx}, {col_idx}): {str(e)}")
        
        print("Cópia concluída com sucesso!")
        
        # Look for the header row (first row with company names)
        header_row = None
        for row_idx in range(1, min(20, max_row)):
            row_values = [new_sheet.cell(row=row_idx, column=col).value for col in range(1, min(20, max_col), 4)]
            non_empty = [val for val in row_values if val is not None]
            if non_empty:
                header_row = row_idx
                print(f"Linha de cabeçalho encontrada: {row_idx}")
                print(f"Valores: {non_empty}")
                break
        
        if header_row is None:
            print("Aviso: Não foi possível encontrar uma linha de cabeçalho.")
            header_row = 1
        
        # Find the most recent date in the file
        print("\nProcurando a data mais recente no arquivo...")
        most_recent_date = None
        date_cells = []
        
        # Check every 4 columns (company data blocks)
        for col_idx in range(1, min(max_col, 2000), 4):
            company_name = new_sheet.cell(row=header_row, column=col_idx).value
            if not company_name:
                continue
                
            print(f"Verificando datas para: {company_name}")
            
            # Look for dates in this column
            for row_idx in range(header_row + 1, min(header_row + 1000, max_row)):
                cell_value = new_sheet.cell(row=row_idx, column=col_idx).value
                
                # Try to identify if it's a date
                date_obj = None
                
                # Check for datetime objects
                if isinstance(cell_value, (datetime.datetime, datetime.date)):
                    date_obj = cell_value
                    print(f"  Data encontrada: {date_obj.strftime('%d/%m/%Y')} na célula ({row_idx}, {col_idx})")
                
                # Check for string dates with / or -
                elif isinstance(cell_value, str) and ('/' in cell_value or '-' in cell_value):
                    try:
                        if '/' in cell_value:
                            parts = cell_value.split('/')
                            if len(parts) == 3:
                                try:
                                    day, month, year = map(int, parts)
                                    # Handle 2-digit years
                                    if year < 100:
                                        year = 2000 + year if year < 50 else 1900 + year
                                    date_obj = datetime.datetime(year, month, day)
                                    print(f"  Data string encontrada: {cell_value} -> {date_obj.strftime('%d/%m/%Y')} na célula ({row_idx}, {col_idx})")
                                except ValueError:
                                    # Try other formats
                                    try:
                                        date_obj = datetime.datetime.strptime(cell_value, "%d/%m/%Y")
                                    except:
                                        pass
                        elif '-' in cell_value:
                            # Try different date formats
                            for fmt in ["%Y-%m-%d", "%d-%m-%Y"]:
                                try:
                                    date_obj = datetime.datetime.strptime(cell_value, fmt)
                                    print(f"  Data string encontrada: {cell_value} -> {date_obj.strftime('%d/%m/%Y')} na célula ({row_idx}, {col_idx})")
                                    break
                                except:
                                    pass
                    except Exception as e:
                        print(f"  Erro ao processar possível data: {cell_value}, erro: {str(e)}")
                
                # Check for Excel serial dates (numbers)
                elif isinstance(cell_value, (int, float)) and 30000 < cell_value < 50000:  # Reasonable range for Excel dates
                    try:
                        # Convert Excel date to Python date
                        date_obj = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(cell_value))
                        print(f"  Data numérica encontrada: {cell_value} -> {date_obj.strftime('%d/%m/%Y')} na célula ({row_idx}, {col_idx})")
                    except Exception as e:
                        print(f"  Erro ao processar possível data numérica: {cell_value}, erro: {str(e)}")
                
                if date_obj:
                    date_cells.append((row_idx, col_idx, date_obj))
                    if most_recent_date is None or date_obj > most_recent_date:
                        most_recent_date = date_obj
                        print(f"  Nova data mais recente: {most_recent_date.strftime('%d/%m/%Y')}")
        
        print(f"Total de datas encontradas: {len(date_cells)}")
        
        # If no dates found, try a more aggressive search
        if not date_cells:
            print("Nenhuma data encontrada no arquivo.")
            print("\nRealizando busca mais abrangente por datas...")
            
            # Scan the entire file for dates
            for col_idx in range(1, min(max_col, 2000)):
                if col_idx % 100 == 0:
                    print(f"  Verificando coluna {col_idx} de {min(max_col, 2000)}...")
                
                for row_idx in range(1, min(max_row, 5000)):
                    cell_value = new_sheet.cell(row=row_idx, column=col_idx).value
                    
                    # Try to identify if it's a date using all methods
                    date_obj = None
                    
                    # Check for datetime objects
                    if isinstance(cell_value, (datetime.datetime, datetime.date)):
                        date_obj = cell_value
                    
                    # Check for string dates
                    elif isinstance(cell_value, str):
                        # Try multiple date formats
                        date_formats = [
                            "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y",
                            "%d-%m-%Y", "%Y-%m-%d", "%m-%d-%Y",
                            "%d.%m.%Y", "%Y.%m.%d", "%m.%d.%Y"
                        ]
                        
                        for fmt in date_formats:
                            try:
                                date_obj = datetime.datetime.strptime(cell_value, fmt)
                                break
                            except:
                                continue
                    
                    # Check for Excel serial dates
                    elif isinstance(cell_value, (int, float)) and 30000 < cell_value < 50000:
                        try:
                            date_obj = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(cell_value))
                        except:
                            pass
                    
                    if date_obj:
                        date_cells.append((row_idx, col_idx, date_obj))
                        if most_recent_date is None or date_obj > most_recent_date:
                            most_recent_date = date_obj
                            print(f"Data encontrada: {date_obj.strftime('%d/%m/%Y')} na célula ({row_idx}, {col_idx})")
            
            print(f"Total de datas encontradas após busca abrangente: {len(date_cells)}")
        
        # If still no dates found, create a JSON with the static data
        if not date_cells:
            print("Ainda não foi possível encontrar datas no arquivo.")
            print("Criando JSON com os dados estáticos sem atualizar datas...")
            
            # Save the static copy anyway
            print(f"\nSalvando arquivo estático em: {output_path}")
            new_wb.save(output_path)
            print("Arquivo estático salvo com sucesso!")
            
            return output_path
        
        # Calculate the date shift needed
        today = datetime.datetime.now().date()
        
        if most_recent_date:
            most_recent_date_only = most_recent_date.date() if isinstance(most_recent_date, datetime.datetime) else most_recent_date
            days_to_shift = (today - most_recent_date_only).days
            print(f"Data mais recente encontrada: {most_recent_date_only}")
            print(f"Dias a avançar: {days_to_shift}")
        else:
            # Default to shifting by 30 days
            days_to_shift = 30
            print(f"Usando deslocamento padrão de {days_to_shift} dias.")
        
        # Update all dates in the file
        print("\nAtualizando datas no arquivo...")
        updated_cells = 0
        
        # First update the cells we already identified
        for row_idx, col_idx, date_obj in date_cells:
            try:
                # Convert date to datetime if needed
                if isinstance(date_obj, datetime.date) and not isinstance(date_obj, datetime.datetime):
                    date_obj = datetime.datetime.combine(date_obj, datetime.time())
                
                new_date = date_obj + datetime.timedelta(days=days_to_shift)
                new_sheet.cell(row=row_idx, column=col_idx).value = new_date
                updated_cells += 1
                
                # Progress indicator
                if updated_cells % 1000 == 0:
                    print(f"  {updated_cells} células atualizadas...")
            except Exception as e:
                print(f"Erro ao atualizar célula ({row_idx}, {col_idx}): {str(e)}")
        
        print(f"Total de células atualizadas: {updated_cells}")
        
        # Save the updated file
        print(f"\nSalvando arquivo atualizado em: {output_path}")
        new_wb.save(output_path)
        print("Arquivo salvo com sucesso!")
        
        return output_path
    
    except Exception as e:
        print(f"Erro ao atualizar datas no arquivo: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

# Modify the process_excel function to include an option to update dates
def process_excel():
    """Process the Excel file and extract stock price history"""
    # Ask for user confirmation
    file_path = download_file()
    if not file_path:
        return
    
    # Inspect the Excel file structure
    inspect_excel_file(file_path)
    
    # Ask if user wants to update dates for testing
    update_dates = input("\nO arquivo parece estar desatualizado. Deseja atualizar as datas para testes? (s/n): ")
    if update_dates.lower() == "s":
        updated_file_path = update_dates_in_file(file_path)
        if updated_file_path:
            use_updated = input("\nDeseja usar o arquivo com datas atualizadas? (s/n): ")
            if use_updated.lower() == "s":
                file_path = updated_file_path
                print(f"Usando arquivo atualizado: {file_path}")
    
    user_input = input("\nDeseja continuar com o processamento do arquivo? (s/n): ")
    if user_input.lower() != "s":
        print("Processamento cancelado pelo usuário.")
        return
    
    # Rest of the function remains the same
    print("\n-- Iniciando processamento do arquivo Excel")
    
    try:
        # Read the Excel file with pandas - try different engines
        print("Lendo arquivo Excel...")
        
        # Try different engines with explicit literals
        df = None
        
        # Try openpyxl first (for .xlsx files)
        try:
            print("Tentando engine: openpyxl")
            df = pd.read_excel(file_path, header=None, engine='openpyxl')
            print("Leitura bem-sucedida com engine: openpyxl")
        except Exception as e:
            print(f"Falha ao ler com engine openpyxl: {str(e)}")
            
            # Try xlrd (for .xls files)
            try:
                print("Tentando engine: xlrd")
                df = pd.read_excel(file_path, header=None, engine='xlrd')
                print("Leitura bem-sucedida com engine: xlrd")
            except Exception as e:
                print(f"Falha ao ler com engine xlrd: {str(e)}")
                
                # Try other engines if needed
                try:
                    print("Tentando engine: odf")
                    df = pd.read_excel(file_path, header=None, engine='odf')
                    print("Leitura bem-sucedida com engine: odf")
                except Exception as e:
                    print(f"Falha ao ler com engine odf: {str(e)}")
        
        if df is None:
            # Try to read as CSV as a fallback
            try:
                print("Tentando ler como CSV...")
                df = pd.read_csv(file_path, header=None)
                print("Leitura bem-sucedida como CSV")
            except Exception as e:
                print(f"Falha ao ler como CSV: {str(e)}")
                raise ValueError("Não foi possível ler o arquivo em nenhum formato suportado")
        
        print(f"Dimensões do DataFrame: {df.shape}")
        
        # Process the data
        historic_data = []
        latest_date = None
        
        # Get the header row
        header_row = df.iloc[0]
        
        # Process data in columns (every 4 columns)
        for col_index in range(0, len(header_row), 4):
            empresa = header_row[col_index]
            codigo = header_row[col_index + 1]
            
            if pd.isna(empresa) or pd.isna(codigo):
                continue
            
            print(f"Processando empresa: {empresa} ({codigo})")
            
            historico_precos = []
            latest_company_date = None
            processed_rows = 0
            
            # Process rows for this company
            for row_index in range(1, len(df)):
                row = df.iloc[row_index]
                
                if col_index >= len(row):
                    continue
                
                data_value = row[col_index]
                preco_value = row[col_index + 1]
                volume_value = row[col_index + 2] if col_index + 2 < len(row) else None
                
                # Skip rows without date or price
                if (pd.isna(data_value) and pd.isna(preco_value)) or pd.isna(data_value):
                    continue
                
                # Process date
                data_formatada = format_date(data_value)
                date_obj = None
                
                try:
                    if data_formatada and "/" in data_formatada:
                        parts = data_formatada.split("/")
                        if len(parts) == 3:
                            day, month, year = map(int, parts)
                            date_obj = datetime.datetime(year, month, day)
                except Exception as e:
                    print(f"Erro ao converter data para objeto: {data_formatada}, {str(e)}")
                
                # Process price
                preco = preco_value
                if pd.isna(preco):
                    continue
                
                if not isinstance(preco, (int, float)):
                    try:
                        preco = float(str(preco).replace(",", "."))
                    except:
                        continue
                
                # Process volume
                volume = volume_value
                if volume is not None:
                    if pd.isna(volume):
                        volume = None
                    elif not isinstance(volume, (int, float)):
                        try:
                            volume = float(str(volume).replace(",", "."))
                        except:
                            volume = None
                
                # Update latest date
                if date_obj:
                    if latest_date is None or date_obj > latest_date:
                        latest_date = date_obj
                    
                    if latest_company_date is None or date_obj > latest_company_date:
                        latest_company_date = date_obj
                
                historico_precos.append({
                    "data": data_formatada,
                    "preco": float(preco),
                    "volume": float(volume) if volume is not None and not pd.isna(volume) else None
                })
                processed_rows += 1
            
            print(f"  Linhas processadas para {codigo}: {processed_rows}")
            
            # Only add companies with valid historical data
            if historico_precos:
                # Sort historical data by date (newest first)
                historico_precos.sort(key=lambda x: datetime.datetime.strptime(x["data"], "%d/%m/%Y") if "/" in x["data"] else datetime.datetime.now(), reverse=True)
                
                latest_date_str = latest_company_date.strftime("%d/%m/%Y") if latest_company_date else "N/A"
                print(f"  Data mais recente para {codigo}: {latest_date_str}")
                
                historic_data.append({
                    "empresa": empresa,
                    "codigo": codigo,
                    "historicoPrecos": historico_precos
                })
        
        if latest_date:
            print(f"Data mais recente encontrada: {latest_date.strftime('%d/%m/%Y')}")
        
        # Save to JSON file
        output_dir = os.path.join(os.path.dirname(__file__), "Finais")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        output_json_path = os.path.join(output_dir, "empresasHistorico.json")
        
        with open(output_json_path, "w", encoding="utf-8") as f:
            json.dump(historic_data, f, ensure_ascii=False, indent=2)
        
        print(f"Historical data JSON file created successfully at {output_json_path}!")
        print(f"Total de empresas processadas: {len(historic_data)}")
    
    except Exception as e:
        print(f"Erro ao processar o arquivo Excel: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    process_excel()