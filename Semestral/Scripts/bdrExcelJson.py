import json
import openpyxl
from openpyxl.utils import get_column_letter
import xlwings as xw
import os
import time
from copy import deepcopy
from datetime import datetime

def reset_file(filename):
    if os.path.exists(filename):
        os.remove(filename)
        
start_time = time.time()

if not os.path.exists('Suporte'):
    os.makedirs('Suporte')
if not os.path.exists('Excel'):
    os.makedirs('Excel')
if not os.path.exists(os.path.join('Suporte', 'Copiar')):
    os.makedirs(os.path.join('Suporte', 'Copiar'))
if not os.path.exists(os.path.join('Suporte', 'Probs')):
    os.makedirs(os.path.join('Suporte', 'Probs'))
    
# Load the JSON data
bdrJson = os.path.join("Finais", "Parcial","bdr.json")
with open(bdrJson, 'r', encoding='utf-8') as f:
    data = json.load(f)

bdrs = data.get("bdrs", [])
bdrs_np = data.get("bdr_nao_patrocinados", [])


# Create a new Excel workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
if ws:
    ws.title = "Precos"

# Initialize row number
row_num = 1
problematic_bdrs = []
problematic_prices = []
problematic_marketcap = []

# Iterate through each company in the JSON data
for bdr in bdrs:
    nomeEmpresa = bdr["nomeEmpresa"]
    codigos = bdr["codigo"]

    # If there are no codes, skip this company and add to no_code_companies list
    if not codigos:
        problematic_bdrs.append(nomeEmpresa)
        continue

    # Iterate through each code and populate the Excel sheet
    if ws:
        ws.cell(row=row_num, column=1, value=nomeEmpresa)
        ws.cell(row=row_num, column=2, value=codigos)  # Garantir que o código seja tratado como string
        ws.cell(row=row_num, column=6, value=f"BVMF:{codigos}")
    row_num += 1

for bdr in bdrs_np:
    nomeEmpresa = bdr["nomeEmpresa"]
    codigos = bdr["codigo"]

    # If there are no codes, skip this company and add to no_code_companies list
    if not codigos:
        problematic_bdrs.append(nomeEmpresa)
        continue

    # Iterate through each code and populate the Excel sheet
    if ws:
        ws.cell(row=row_num, column=1, value=nomeEmpresa)
        ws.cell(row=row_num, column=2, value=codigos)  # Garantir que o código seja tratado como string
        ws.cell(row=row_num, column=6, value=f"BVMF:{codigos}")
    row_num += 1
        
# Ensure the file is not open or remove it if it already exists
filename = os.path.join("Excel", "precoBdr.xlsx")
if os.path.exists(filename):
    os.remove(filename)

# Save the workbook
wb.save(filename)
print("precoBdr.xlsx criado")

# Open the workbook with xlwings to adjust columns and add macro
app = xw.App(visible=False)
wb_xw = app.books.open(filename)
ws_xw = wb_xw.sheets["Precos"]

# Add the VBA macro
vba_code = '''
Sub ConvertToStock()
    Set rng = Columns("F:F")
    On Error Resume Next
    Set rng = rng.SpecialCells(xlCellTypeConstants)
    On Error GoTo 0
    If Not rng Is Nothing Then
        rng.ConvertToLinkedDataType ServiceID:=268435456, LanguageCulture:= "en-US"
    Else
        MsgBox "Não há células com valores na coluna F."
    End If
End Sub
'''

# Add the macro to the workbook
wb_xw.api.VBProject.VBComponents.Add(1).CodeModule.AddFromString(vba_code)

# Save the workbook with the macro
macro_filename = "precoBdr_with_macro.xlsm"
wb_xw.save(macro_filename)
wb_xw.close()

# Reopen the workbook to run the macro
wb_macro = app.books.open(macro_filename)
time.sleep(5)  # Adding delay to ensure Excel is ready

# Run the macro
try:
    app.macro('ConvertToStock')()
except Exception as e:
    print(f"Erro ao executar a macro: {e}")
    wb_macro.close()
    app.quit()
    exit()

# Reopen the workbook and populate columns 3 and 4 again with the formulas
ws_macro = wb_macro.sheets["Precos"]
for row in range(1, row_num):
    ws_macro.cells(row, 3).value = f"=F{row}.[Price]"
    ws_macro.cells(row, 4).value = f"=F{row}.[Market cap]"

ws_macro.range('A:F').columns.autofit()
wb_macro.save(filename)
time.sleep(1)

# Reopen workbook to check values
wb_macro = app.books.open(filename)
time.sleep(5)  

# Check for problems in columns 3 and 4
rows_to_delete = []
for row in range(1, row_num):
    price_cell = ws_macro.cells(row, 3).value
    market_cap_cell = ws_macro.cells(row, 4).value
    company_code = ws_macro.cells(row, 2).value
    if price_cell == -2146826239 and market_cap_cell == -2146826239:
        problematic_bdrs.append(company_code)
        rows_to_delete.append(row)
    elif price_cell == 0.0 or isinstance(price_cell, str):
        problematic_prices.append(company_code)
        rows_to_delete.append(row)
    elif market_cap_cell == -2146826239:
        problematic_marketcap.append(company_code)

# Delete problematic rows from the Excel sheet
for row in sorted(rows_to_delete, reverse=True):
    ws_macro.range(f"A{row}:F{row}").delete()

wb_macro.save(filename)
wb_macro.close()
app.quit()

# Update the JSON data with problematic companies
def add_problematic_company(companies, problematic_companies, code):
    company_to_add = None
    for company in companies:
        if company["codigo"] == code:  # Direct comparison since codigo is a single value
            companies.remove(company)
            company_to_add = company
            break
    
    if company_to_add:
        # Verificar se problematic_companies contém strings ou dicionários
        if problematic_companies and isinstance(problematic_companies[0], dict):
            # Se for lista de dicionários, verifica pelo codigoCVM
            for existing_company in problematic_companies:
                if existing_company["codigoCVM"] == company_to_add["codigoCVM"]:
                    return
        else:
            # Se for lista de strings, verifica pelo código
            if code in problematic_companies:
                return
        problematic_companies.append(company_to_add)

# Handle problematic companies
for code in problematic_bdrs + problematic_prices:
    add_problematic_company(bdrs, problematic_bdrs, code)

# Update the JSON structure
data["bdrs"] = bdrs
data["bdr_nao_patrocinados"] = bdrs_np

# Save the updated JSON file
with open(bdrJson, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

codes = []
txt_Codigos = os.path.join("Finais", "Copiar", "codigosBdr.txt")
reset_file(txt_Codigos)

with open(txt_Codigos, "w", encoding="utf-8") as f:
    for company in bdrs:
        codigo = company["codigo"]
        f.write(f"{codigo}\n")
        codes.append(codigo)
    for company in bdrs_np:
        codigo = company["codigo"]
        f.write(f"{codigo}\n")
        codes.append(codigo)
         
# Get absolute paths
current_dir = os.getcwd()
empresaHistorico = os.path.join(current_dir, "Excel", "historicoBdr.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
if ws:
    ws.title = "Historico"
col_num = 1

for i, codigo in enumerate(codes):
    col_data = col_num + 2
    col_letter = get_column_letter(col_num)
    col_data_letter = get_column_letter(col_data)
    if ws:
        ws.cell(row=1, column=col_num, value=codigo)
        ws.cell(row=1, column=col_data, value="10/01/1994")
        formula = f"HISTÓRICODEAÇÕES({col_letter}1,{col_data_letter}1,HOJE(),0,0,0,1,5)"
        ws.cell(row=2, column=col_num).value = f"={formula}"
    col_num += 4
    
# Make sure the file doesn't exist before saving
if os.path.exists(empresaHistorico):
    try:
        os.remove(empresaHistorico)
    except:
        print(f"Não foi possível remover o arquivo existente: {empresaHistorico}")
        # Generate a unique filename if we can't delete the existing one
        empresaHistorico = os.path.join(current_dir, "Excel", f"historicoBdr_{int(time.time())}.xlsx")

wb.save(empresaHistorico)
print(f"historicoBdr.xlsx criado em: {empresaHistorico}")
wb.close()
time.sleep(5)

# Ensure Excel is fully closed before proceeding
try:
    for proc in os.popen('tasklist').readlines():
        if 'EXCEL.EXE' in proc:
            pid = int(proc.split()[1])
            os.system(f'taskkill /F /PID {pid}')
            time.sleep(2)
            break
except:
    print("Não foi possível verificar se o Excel está em execução")

# Abrir o arquivo com xlwings para manipular as fórmulas
app = xw.App(visible=False)
try:
    print(f"Tentando abrir: {empresaHistorico}")
    wb_historico = app.books.open(empresaHistorico)
    ws2 = wb_historico.sheets["Historico"]

    print("Ajustando fórmulas...")
    used_range = ws2.range("A1").expand()
    for cell in used_range:
        if cell.formula and cell.formula.startswith("=@HISTÓRICODEAÇÕES"):
            cell.formula = cell.formula.replace("=@HISTÓRICODEAÇÕES", "=HISTÓRICODEAÇÕES")

    time.sleep(5)
    wb_historico.save()
    wb_historico.close()
    print("Arquivo histórico processado com sucesso")
except Exception as e:
    print(f"Erro ao processar o arquivo histórico: {e}")
    print(f"Caminho do arquivo: {empresaHistorico}")
    print("Continuando com o resto do processo...")
finally:
    try:
        app.quit()
        # Ensure Excel is fully closed
        for proc in os.popen('tasklist').readlines():
            if 'EXCEL.EXE' in proc:
                pid = int(proc.split()[1])
                os.system(f'taskkill /F /PID {pid}')
                break
    except:
        pass

print("Fórmulas ajustadas em historicoBdr.xlsx")
             
end_time = time.time()
execution_time_min = (end_time - start_time)/60
execution_time_secs = (end_time - start_time)%60
print(f"Tempo de execução: {execution_time_min:.0f} minutos e {execution_time_secs:.0f} segundos")
    
# Create a text file with problematic companies
txt_filename = os.path.join("Suporte", "Probs", "problemaBdr.txt")
reset_file(txt_filename)
len_bdr = len(bdrs) + len(bdrs_np)
with open(txt_filename, "w", encoding="utf-8") as f:
    f.write(f"Relatório gerado em: {datetime.now().strftime('%d / %m / %Y  %H:%M')}\n\n")
    f.write(f"Tempo de execução: {execution_time_min:.0f} minutos e {execution_time_secs:.0f} segundos.\n\n")
    f.write(f"BDRs: {len_bdr}\n\n")
    f.write(f"BDRs com problema: {len(problematic_bdrs)}\n\n")
    if problematic_bdrs:
        f.write(f"\nBDRs problemáticas:\n")
        for company in problematic_bdrs:
            f.write(f"{company}\n")
    if problematic_prices:
        f.write(f"\nBDRs com problema preço:\n")        
        for company in problematic_prices:    
            f.write(f"{company}\n")
    if problematic_marketcap:
        f.write(f"\nBDRs com problema valor de mercado:\n")        
        for company in problematic_marketcap:    
            f.write(f"{company}\n")

reset_file('precoBdr_with_macro.xlsm')
print("Verificação realizada: 'problemaBdr.txt'")
print("Codigos: 'codigosBdr.txt'")
