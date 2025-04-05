import json
import openpyxl
from openpyxl.utils import get_column_letter
import xlwings as xw
import os
import time
from copy import deepcopy
from datetime import datetime

def reset_file(filename):
    if os.path.exists(filename):
        os.remove(filename)
        
start_time = time.time()

# Create necessary directories
if not os.path.exists('Suporte'):
    os.makedirs('Suporte')
if not os.path.exists(os.path.join('Suporte', 'Copiar')):
    os.makedirs(os.path.join('Suporte', 'Copiar'))
if not os.path.exists(os.path.join('Suporte', 'Probs')):
    os.makedirs(os.path.join('Suporte', 'Probs'))
if not os.path.exists('Excel'):
    os.makedirs('Excel')
if not os.path.exists('Finais'):
    os.makedirs('Finais')
    
# Load the JSON data
empresasJson = os.path.join("Finais/Parcial", "empresasParcial.json")
with open(empresasJson, 'r', encoding='utf-8') as f:
    data = json.load(f)

companies = data.get("empresas", [])
companies_no_code = data.get("empresas_sem_codigo", [])
problematic_companies_data = data.get("empresas_problema", [])

# Create a new Excel workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
if ws:
    ws.title = "Precos"

# Initialize row number
row_num = 1
no_code_companies = []
problematic_companies = []
problematic_prices = []
problematic_marketcap = []

# Iterate through each company in the JSON data
for company in companies:
    nomeEmpresa = company["nomeEmpresa"]
    codigos = company["codigos"]

    # If there are no codes, skip this company and add to no_code_companies list
    if not codigos:
        no_code_companies.append(nomeEmpresa)
        continue

    # Iterate through each code and populate the Excel sheet
    for i, codigo in enumerate(codigos, start=1):
        if ws:
            ws.cell(row=row_num, column=1, value=nomeEmpresa)
            ws.cell(row=row_num, column=2, value=codigo)
            ws.cell(row=row_num, column=6, value=f"BVMF:{codigo}")
        row_num += 1

# Ensure the file is not open or remove it if it already exists

filename = os.path.join("Excel", "precoEmpresa.xlsx")
if os.path.exists(filename):
    os.remove(filename)

# Save the workbook
wb.save(filename)
print("precoEmpresa.xlsx criado")

# Open the workbook with xlwings to adjust columns and add macro
app = xw.App(visible=False)
wb_xw = app.books.open(filename)
ws_xw = wb_xw.sheets["Precos"]

# Add the VBA macro
vba_code = '''
Sub ConvertToStock()
    Columns("F:F").Select
    Selection.ConvertToLinkedDataType ServiceID:=268435456, LanguageCulture:= "en-US"
End Sub
'''

# Add the macro to the workbook
wb_xw.api.VBProject.VBComponents.Add(1).CodeModule.AddFromString(vba_code)

# Save the workbook with the macro
macro_filename = "precoEmpresa_with_macro.xlsm"
wb_xw.save(macro_filename)
wb_xw.close()

# Reopen the workbook to run the macro
wb_macro = app.books.open(macro_filename)
time.sleep(5)  # Adding delay to ensure Excel is ready

# Run the macro
try:
    app.macro('ConvertToStock')()
except Exception as e:
    print(f"Erro ao executar a macro: {e}")
    wb_macro.close()
    app.quit()
    exit()

# Reopen the workbook and populate columns 3 and 4 again with the formulas
ws_macro = wb_macro.sheets["Precos"]
for row in range(1, row_num):
    ws_macro.cells(row, 3).value = f"=F{row}.[Price]"
    ws_macro.cells(row, 4).value = f"=F{row}.[Market cap]"

# Auto-fit columns A:F
ws_macro.range('A:F').columns.autofit()

# Save the workbook as a normal Excel file
wb_macro.save(filename)

# Wait for Excel to close
time.sleep(1)

# Reopen workbook to check values
wb_macro = app.books.open(filename)
time.sleep(5)  # Adding delay to ensure Excel is ready

# Check for problems in columns 3 and 4
rows_to_delete = []
for row in range(1, row_num):
    price_cell = ws_macro.cells(row, 3).value
    market_cap_cell = ws_macro.cells(row, 4).value
    company_code = ws_macro.cells(row, 2).value
    if price_cell == -2146826239 and market_cap_cell == -2146826239:
        problematic_companies.append(company_code)
        rows_to_delete.append(row)
    elif price_cell == 0.0 or isinstance(price_cell, str) or price_cell == "#CAMPO!":
        problematic_prices.append(company_code)
        rows_to_delete.append(row)
    elif market_cap_cell == -2146826239:
        problematic_marketcap.append(company_code)

# Delete problematic rows from the Excel sheet
for row in sorted(rows_to_delete, reverse=True):
    ws_macro.range(f"A{row}:F{row}").delete()

# Save the updated workbook
wb_macro.save(filename)
wb_macro.close()
app.quit()

# Update the JSON data with problematic companies
def add_problematic_company(companies, problematic_companies, code):
    company_to_add = None
    for company in companies:
        if code in company["codigos"]:
            if len(company["codigos"]) == 1:
                companies.remove(company)
                company_to_add = company
            else:
                company["codigos"].remove(code)
                company_to_add = deepcopy(company)
                company_to_add["codigos"] = [code]
            break
    if company_to_add:
        for existing_company in problematic_companies:
            if existing_company["codigoCVM"] == company_to_add["codigoCVM"]:
                existing_company["codigos"].append(code)
                return
        problematic_companies.append(company_to_add)

# Handle problematic companies
for code in problematic_companies + problematic_prices + problematic_marketcap:
    add_problematic_company(companies, problematic_companies_data, code)

# Update the JSON structure
data["empresas"] = companies
data["empresas_problema"] = problematic_companies_data
data["empresas_sem_codigo"] = companies_no_code

# Save the updated JSON file
with open(empresasJson, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

empresaFinalJson = os.path.join("Finais/Parcial", "empresas.json")
with open(empresaFinalJson, 'w', encoding='utf-8') as f:
    json.dump(companies, f, ensure_ascii=False, indent=4)

codes = []
txt_Codigos = os.path.join("Finais", "Copiar", "codigosEmpresa.txt")
reset_file(txt_Codigos)

with open(txt_Codigos, "w", encoding="utf-8") as f:
    for company in companies:
        for codigo in company["codigos"]:
            f.write(f"{codigo}\n")
            codes.append(codigo)

         
wb = openpyxl.Workbook()
ws = wb.active
if ws:
    ws.title = "Historico"
col_num = 1

for i, codigo in enumerate(codes):
    col_data = col_num + 2
    col_letter = get_column_letter(col_num)
    col_data_letter = get_column_letter(col_data)
    if ws:
        ws.cell(row=1, column=col_num, value=codigo)
        ws.cell(row=1, column=col_data, value="10/01/1950")
        formula = f"HISTÓRICODEAÇÕES({col_letter}1,{col_data_letter}1,HOJE(),0,0,0,1,5)"
        # formula = f"HISTÓRICODEAÇÕES({col_letter}1,10/01/1950,HOJE(),0,0,0,1,5)"
        ws.cell(row=2, column=col_num).value = f"={formula}"
    col_num += 4
    
empresaHistorico = os.path.join("Excel", "historicoEmpresa.xlsx")      
wb.save(empresaHistorico)
print("historicoEmpresa.xlsx criado")
wb.close()
time.sleep(5)
# Abrir o arquivo com xlwings para manipular as fórmulas
app = xw.App(visible=False)
wb_historico = app.books.open(empresaHistorico)
ws2 = wb_historico.sheets["Historico"]

# Iterar sobre as células para ajustar as fórmulas
used_range = ws2.range("A1").expand()
for cell in used_range:
    if cell.formula and cell.formula.startswith("=@HISTÓRICODEAÇÕES"):
        cell.formula = cell.formula.replace("=@HISTÓRICODEAÇÕES", "=HISTÓRICODEAÇÕES")

time.sleep(5)
# Salvar as alterações
wb_historico.save()
wb_historico.close()
app.quit()

print("Fórmulas ajustadas em historicoEmpresa.xlsx")
             
end_time = time.time()
execution_time_min = (end_time - start_time)/60
execution_time_secs = (end_time - start_time)%60
print(f"Tempo de execução: {execution_time_min:.0f} minutos e {execution_time_secs:.0f} segundos")
    
# Create a text file with problematic companies
txt_filename = os.path.join("Suporte", "Probs", "problemaEmpresa.txt")
reset_file(txt_filename)
with open(txt_filename, "w", encoding="utf-8") as f:
    f.write(f"Relatório gerado em: {datetime.now().strftime('%d / %m / %Y  %H:%M')}\n\n")
    f.write(f"Tempo de execução: {execution_time_min:.0f} minutos e {execution_time_secs:.0f} segundos.\n\n")
    f.write(f"Companhias: {len(companies)}\n\n")
    f.write(f"Companhias com problema: {len(problematic_companies_data)}\n\n")
    f.write(f"Companhias sem codigo: {len(companies_no_code)}\n\n")
    if no_code_companies:
        f.write(f"Empresas sem código:\n")
        for company in no_code_companies:
            f.write(f"{company}\n")
    if problematic_companies:
        f.write(f"\nCódigos problemáticos:\n")
        for company in problematic_companies:
            f.write(f"{company}\n")
    if problematic_prices:
        f.write(f"\nCódigos com problema preço:\n")        
        for company in problematic_prices:    
            f.write(f"{company}\n")
    if problematic_marketcap:
        f.write(f"\nCódigos com problema valor de mercado:\n")        
        for company in problematic_marketcap:    
            f.write(f"{company}\n")

        
reset_file('precoEmpresa_with_macro.xlsm')
print("Verificação realizada: 'problemaEmpresa.txt'")
print("Codigos: 'codigosEmpresa.txt'")
print(f"{len(companies)} empresas com códigos funcionais.'")
