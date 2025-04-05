import json
import openpyxl
from openpyxl.utils import get_column_letter
import xlwings as xw
import os
import time
from datetime import datetime

# Função para resetar um arquivo se ele existir
def reset_file(filename):
    if os.path.exists(filename):
        os.remove(filename)

start_time = time.time()

# Verifica se a pasta "Suporte" existe, caso contrário, cria-a
if not os.path.exists('Suporte'):
    os.makedirs('Suporte')

# Caminho para o arquivo JSON inicial dos FIIs
fiisJson = os.path.join("Finais", "Parcial", "fiis.json")

# Leitura dos dados do JSON
with open(fiisJson, 'r', encoding='utf-8') as f:
    data = json.load(f)

# Lista de FIIs com códigos, sem códigos e com problemas
fiis_with_codes = data
problematic_fiis_data = []

# Criar um novo workbook Excel e selecionar a planilha ativa para os preços
wb = openpyxl.Workbook()
ws_preco = wb.active
if ws_preco:
    ws_preco.title = "Preços"

# Inicializar número da linha
row_num = 1
no_code_fiis = []
problematic_fiis = []

# Iterar através de cada FII nos dados JSON
for fii in fiis_with_codes:
    nomeFII = fii["nomeFII"]
    codigos = fii["codigo"]

    # Se não houver código, adicionar à lista de FIIs sem código
    if not codigos:
        no_code_fiis.append(nomeFII)
        continue

    # Se codigos for uma lista, iterar sobre ela
    if isinstance(codigos, list):
        for codigo in codigos:
            # Preencher a planilha do Excel com o código do FII
            if ws_preco:
                ws_preco.cell(row=row_num, column=1, value=nomeFII)
                ws_preco.cell(row=row_num, column=2, value=codigo)
                ws_preco.cell(row=row_num, column=6, value=f"BVMF:{codigo}")
            row_num += 1
    else:
        # Se não for uma lista, tratar como string única
        if ws_preco:
            ws_preco.cell(row=row_num, column=1, value=nomeFII)
            ws_preco.cell(row=row_num, column=2, value=codigos)
            ws_preco.cell(row=row_num, column=6, value=f"BVMF:{codigos}")
        row_num += 1

# Garantir que o arquivo não esteja aberto ou removê-lo se já existir
preco_filename = os.path.join("Excel", "precoFII.xlsx")
if os.path.exists(preco_filename):
    os.remove(preco_filename)

# Salvar o workbook Excel
wb.save(preco_filename)
print("precoFII.xlsx criado")

# Abrir o workbook com xlwings para ajustar as colunas e adicionar a macro
app = xw.App(visible=False)
wb_xw = app.books.open(preco_filename)
ws_xw = wb_xw.sheets["Preços"]

# Adicionar a macro VBA
vba_code = '''
Sub ConvertToStock()
    Columns("F:F").Select
    Selection.ConvertToLinkedDataType ServiceID:=268435456, LanguageCulture:= "en-US"
End Sub
'''

# Adicionar a macro ao workbook
wb_xw.api.VBProject.VBComponents.Add(1).CodeModule.AddFromString(vba_code)

# Salvar o workbook com a macro
macro_filename = "precoFII_with_macro.xlsm"
wb_xw.save(macro_filename)
wb_xw.close()

# Reabrir o workbook para executar a macro
wb_macro = app.books.open(macro_filename)
time.sleep(5)  # Adicionar um pequeno delay para garantir que o Excel esteja pronto

# Executar a macro
try:
    app.macro('ConvertToStock')()
except Exception as e:
    print(f"Erro ao executar a macro: {e}")
    wb_macro.close()
    app.quit()
    exit()

# Reabrir o workbook e preencher novamente as colunas 3 e 4 com as fórmulas
ws_macro = wb_macro.sheets["Preços"]
for row in range(1, row_num):
    ws_macro.cells(row, 3).value = f"=F{row}.[Price]"

# Ajustar automaticamente as colunas A:F
ws_macro.range('A:F').columns.autofit()

# Salvar o workbook como um arquivo Excel normal
wb_macro.save(preco_filename)

# Aguardar o Excel fechar
time.sleep(1)

# Reabrir o workbook para verificar os valores
wb_macro = app.books.open(preco_filename)
time.sleep(5)  # Adicionar um pequeno delay para garantir que o Excel esteja pronto

# Verificar problemas nas colunas 3 e 4
rows_to_delete = []
for row in range(1, row_num):
    price_cell = ws_macro.cells(row, 3).value
    market_cap_cell = ws_macro.cells(row, 4).value
    fii_code = ws_macro.cells(row, 2).value
    if price_cell == -2146826239:
        problematic_fiis.append(fii_code)
        rows_to_delete.append(row)
    elif price_cell == 0.0 or isinstance(price_cell, str) or price_cell == "#CAMPO!":
        problematic_fiis.append(fii_code)
        rows_to_delete.append(row)

# Deletar as linhas problemáticas da planilha Excel
for row in sorted(rows_to_delete, reverse=True):
    ws_macro.range(f"A{row}:F{row}").delete()

# Salvar o workbook atualizado
wb_macro.save(preco_filename)
wb_macro.close()
app.quit()

# Função para adicionar um FII problemático aos dados problemáticos
def add_problematic_fii(fiis, problematic_fiis, code):
    for fii in fiis:
        if code in fii["codigo"]:
            fii["codigo"].remove(code)
            if not fii["codigo"]:
                problematic_fiis.append(fii)
                fiis.remove(fii)
            break

# Lidar com os FIIs problemáticos
for code in problematic_fiis:
    add_problematic_fii(fiis_with_codes, problematic_fiis_data, code)

# Atualizar a estrutura JSON
data = fiis_with_codes

# Salvar o arquivo JSON atualizado
with open(fiisJson, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

# Criar um arquivo de texto com todos os códigos de FIIs
codes = []
txt_Codigos = os.path.join("Finais", "Copiar", "codigosFII.txt")
reset_file(txt_Codigos)

with open(txt_Codigos, "w", encoding="utf-8") as f:
    for fii in data:
        for codigo in fii['codigo']:
            if codigo.strip():
                f.write(f"{codigo}\n")
                codes.append(codigo)

# Criar um workbook Excel para o histórico
wb_historico = openpyxl.Workbook()
ws_historico = wb_historico.active
if ws_historico:
    ws_historico.title = "Histórico"
col_num = 1

# Iterar sobre os códigos para preencher as fórmulas
for i, codigo in enumerate(codes):
    col_data = col_num + 2
    col_letter = get_column_letter(col_num)
    col_data_letter = get_column_letter(col_data)
    if ws_historico:
        ws_historico.cell(row=1, column=col_num, value=codigo)
        # Fórmula de exemplo para o histórico de FIIs
        ws_historico.cell(row=1, column=col_data, value="10/01/1950")
        formula = f"HISTÓRICODEAÇÕES({col_letter}1,{col_data_letter}1,HOJE(),0,0,0,1,5)"
        # formula = f"HISTÓRICODEAÇÕES({col_letter}1,'10/01/1950',HOJE(),0,0,0,1,5)"
        ws_historico.cell(row=2, column=col_num).value = f"={formula}"
    col_num += 4

# Caminho para o arquivo Excel final do histórico
fiiHistorico = os.path.join("Excel", "historicoFII.xlsx")
wb_historico.save(fiiHistorico)
print("historicoFII.xlsx criado")
wb_historico.close()
time.sleep(5)

# Abrir o arquivo com xlwings para ajustar as fórmulas
app = xw.App(visible=False)
wb_historico = app.books.open(fiiHistorico)
ws2 = wb_historico.sheets["Histórico"]

# Iterar sobre as células para ajustar as fórmulas
used_range = ws2.range("A1").expand()
for cell in used_range:
    if cell.formula and cell.formula.startswith("=@HISTÓRICODEAÇÕES"):
        cell.formula = cell.formula.replace("=@HISTÓRICODEAÇÕES", "=HISTÓRICODEAÇÕES")

time.sleep(5)
# Salvar as alterações
wb_historico.save()
wb_historico.close()
app.quit()

# Criar um arquivo de texto com os FIIs problemáticos
txt_filename = os.path.join("Suporte", "Probs", "problemaFII.txt")
reset_file(txt_filename)
with open(txt_filename, "w", encoding="utf-8") as f:
    f.write(f"Relatório gerado em: {datetime.now().strftime('%d / %m / %Y  %H:%M')}\n\n")
    f.write(f"Tempo de execução: {(time.time() - start_time)/60:.0f} minutos.\n\n")
    f.write(f"FIIs: {len(fiis_with_codes)}\n\n")
    f.write(f"FIIs com problema: {len(problematic_fiis_data)}\n\n")
    f.write(f"FIIs sem código: {len(no_code_fiis)}\n\n")
    if no_code_fiis:
        f.write(f"FIIs sem código:\n")
        for fii in no_code_fiis:
            f.write(f"{fii}\n")
    if problematic_fiis:
        f.write(f"\nCódigos problemáticos:\n")
        for fii in problematic_fiis:
            f.write(f"{fii}\n")

print("Verificação concluída: 'problemaFII.txt'")
print("Códigos salvos em: 'codigosFII.txt'")
print(f"{len(fiis_with_codes)} FIIs com códigos funcionais.")

end_time = time.time()
execution_time_min = (end_time - start_time) / 60
execution_time_secs = (end_time - start_time) % 60
print(f"Tempo de execução: {execution_time_min:.0f} minutos e {execution_time_secs:.0f} segundos")
